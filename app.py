"""
Telegram Recall Bot
--------------------
Webhook-based Telegram bot. Message it your grocery list (one item per line)
and it replies with active FDA/USDA recalls filtered to California/nationwide.

Deploy this as a Render free web service. Render's free web services spin
down after 15 min idle and cold-start on the next request (~30-50s) — that's
fine here since Telegram retries delivery, you'll just get a slightly slower
first reply after a quiet period.
"""

import os
import re
import io
import time
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from flask import Flask, request, render_template_string

app = Flask(__name__)

BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_API = f"https://api.telegram.org/bot{BOT_TOKEN}"

FDA_URL = "https://api.fda.gov/food/enforcement.json"
USDA_URL = "https://www.fsis.usda.gov/fsis/api/recall/v/1"


# ---------- recall lookup (same logic as the standalone script) ----------

def location_matches(text):
    if not text:
        return False
    t = text.lower()
    return "california" in t or "nationwide" in t


def check_fda(term):
    matches = _check_fda_query(term)
    if not matches:
        # try the singular/plural stem too (e.g. "blueberries" -> "blueberr")
        # as a second safe, quoted-phrase query — no wildcard syntax risk
        stem = word_stem(term)
        if stem and stem != term.lower():
            matches = _check_fda_query(stem)
    return matches


def _check_fda_query(term):
    params = {
        "search": f'product_description:"{term}"+AND+status:"Ongoing"',
        "sort": "report_date:desc",
        "limit": 25,
    }
    try:
        r = requests.get(FDA_URL, params=params, timeout=15)
        if r.status_code == 404:
            return []
        r.raise_for_status()
        results = r.json().get("results", [])
        matches = [x for x in results if location_matches(x.get("distribution_pattern"))]
        return [
            {
                "source": "FDA",
                "brand": m.get("recalling_firm", "Unknown brand"),
                "product": (m.get("product_description") or "")[:90],
                "reason": m.get("reason_for_recall", "No reason listed"),
                "classification": m.get("classification", ""),
                "date": format_date_str(m.get("report_date", "")),
                "terminated": "unknown",
            }
            for m in matches
        ]
    except requests.RequestException:
        return []


def check_usda(term):
    params = {"field_product_items_value": term, "field_active_notice": "True"}
    try:
        r = requests.get(USDA_URL, params=params, headers=HEADERS, timeout=15)
        r.raise_for_status()
        results = r.json()
        if not isinstance(results, list):
            return []
        matches = [x for x in results if location_matches(x.get("field_states"))]
        return [
            {
                "source": "USDA",
                "brand": m.get("field_establishment", "Unknown establishment"),
                "product": (m.get("field_title") or "")[:100],
                "reason": m.get("field_recall_reason", "No reason listed"),
                "classification": m.get("field_recall_classification", ""),
                "date": format_date_str(m.get("field_recall_date", "")),
                "terminated": "unknown",
            }
            for m in matches
        ]
    except requests.RequestException:
        return []


# ---------- press-release scraping (catches very recent recalls the ----------
# ---------- classification-lagged APIs above haven't picked up yet)  ----------

FDA_PRESS_URL = "https://www.fda.gov/safety/recalls-market-withdrawals-safety-alerts"
FSIS_PRESS_URL = "https://www.fsis.usda.gov/recalls"

HEADERS = {"User-Agent": "Mozilla/5.0 (grocery-recall-bot/1.0)"}


def word_stem(word):
    """Crude singular/plural stem so 'blueberries' still matches text
    saying 'blueberry', and vice versa."""
    w = word.lower()
    if w.endswith("ies") and len(w) > 4:
        return w[:-3]
    if w.endswith("es") and len(w) > 3:
        return w[:-2]
    if w.endswith("s") and len(w) > 3:
        return w[:-1]
    return w


def term_matches(term, haystack):
    """True if the term (or its singular/plural stem) appears in haystack."""
    h = haystack.lower()
    t = term.lower().strip()
    if t in h:
        return True
    stem = word_stem(t)
    return bool(stem) and stem in h


FDA_XLSX_URL = "https://www.fda.gov/safety/recalls-market-withdrawals-safety-alerts/datatables-data?_format=xlsx"


def check_fda_press(term):
    """Fetch FDA's full press-release recall dataset as XLSX (the same
    file behind the page's 'Download XLSX' link) and search it directly.
    Replaces an earlier page-by-page HTML scrape: testing showed the
    page's ?page= parameter is ignored (the table pagination is
    client-side JavaScript), so every 'page' silently returned identical
    content. The XLSX export isn't paginated at all — one request gets
    the whole dataset."""
    try:
        r = requests.get(FDA_XLSX_URL, headers=HEADERS, timeout=30)
        r.raise_for_status()
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        header = [str(h).strip().lower().replace("-", " ") if h else "" for h in header]

        def col_idx(fragment):
            for i, h in enumerate(header):
                if fragment in h:
                    return i
            return None

        idx = {
            "date": col_idx("date"),
            "brand": col_idx("brand"),
            "product": col_idx("product description"),
            "type": col_idx("product type"),
            "reason": col_idx("reason"),
            "company": col_idx("company"),
            "terminated": col_idx("terminated"),
        }

        def cell(row, key):
            i = idx.get(key)
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        matches = []
        for row in rows:
            ptype = cell(row, "type")
            if "food" not in ptype.lower():
                continue
            brand = cell(row, "brand")
            product = cell(row, "product")
            company = cell(row, "company")
            reason = cell(row, "reason")
            date = cell(row, "date")
            terminated_raw = cell(row, "terminated")
            terminated = "Yes" if terminated_raw.strip().lower() in ("yes", "y", "true") else "No"
            haystack = f"{brand} {product} {company}"
            if term_matches(term, haystack):
                matches.append({
                    "source": "FDA press release",
                    "brand": company or brand,
                    "product": product,
                    "reason": reason,
                    "classification": "",
                    "date": date,
                    "terminated": terminated,
                })
        return matches
    except Exception:
        return []


def check_fsis_press(term):
    """Scrape FSIS's recall/alert list page the same way, for meat/poultry/egg."""
    try:
        r = requests.get(FSIS_PRESS_URL, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        matches = []
        # FSIS lists recalls as headline links followed by a summary paragraph
        for link in soup.find_all("a"):
            headline = link.get_text(" ", strip=True)
            if not headline or len(headline) < 15:
                continue
            summary = ""
            sibling = link.find_next("p")
            if sibling:
                summary = sibling.get_text(" ", strip=True)
            haystack = f"{headline} {summary}"
            if term_matches(term, haystack):
                matches.append({
                    "source": "USDA press release",
                    "brand": headline,
                    "product": "",
                    "reason": summary[:150],
                    "classification": "",
                    "date": "",
                    "terminated": "unknown",
                })
        # de-dupe by headline, cap results
        seen = set()
        deduped = []
        for m in matches:
            if m["brand"] not in seen:
                seen.add(m["brand"])
                deduped.append(m)
        return deduped[:5]
    except requests.RequestException:
        return []


def format_date_str(value):
    """Normalize dates from different sources into MM/DD/YYYY."""
    if not value:
        return "date unknown"
    v = str(value).strip()
    if len(v) == 8 and v.isdigit():  # FDA API's raw YYYYMMDD
        return f"{v[4:6]}/{v[6:8]}/{v[0:4]}"
    return v


def normalize_brand(name):
    """Loosely normalize a company/brand name so the same real-world
    recall is recognized as a duplicate even when different sources word
    it slightly differently (e.g. 'The Hampton Grocers, Inc.' vs
    'The Hampton Grocer, Inc.')."""
    n = (name or "").lower()
    n = re.sub(r"[^a-z0-9\s]", " ", n)
    n = re.sub(
        r"\b(the|inc|llc|co|corp|company|ltd|group|foods?|farms?)\b", " ", n
    )
    n = re.sub(r"s\b", "", n)  # crude plural/possessive trim: grocers -> grocer
    return re.sub(r"\s+", " ", n).strip()


def check_item(term):
    term = term.strip()
    if not term:
        return []
    return (
        check_fda(term)
        + check_usda(term)
        + check_fda_press(term)
        + check_fsis_press(term)
    )


# ---------- parsing a pasted grocery list into item terms ----------

BULLET_RE = re.compile(r"^[\s•\-\*\u2022]+")
QTY_RE = re.compile(r"\(\d+\)\s*$")


def parse_list(text):
    items = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = BULLET_RE.sub("", line).strip()
        line = QTY_RE.sub("", line).strip()
        if not line:
            continue
        # skip section headers / store lines: ALL CAPS, or a slash-joined store list
        letters = [c for c in line if c.isalpha()]
        if letters and all(c.isupper() for c in letters):
            continue
        if "/" in line and " " not in line:
            continue
        items.append(line)
    return items


def build_reply(items):
    if not items:
        return "Didn't find any grocery items in that message — send one item per line."

    header = "🛒 Recall check (California / nationwide):"
    item_blocks = []
    any_recalls = False

    for item in items:
        matches = check_item(item)
        # de-dupe by normalized brand name — different sources word the same
        # company slightly differently, so exact-text matching under-deduped
        seen_keys = set()
        deduped = []
        for m in matches:
            key = normalize_brand(m["brand"])
            if key not in seen_keys:
                seen_keys.add(key)
                deduped.append(m)
        matches = deduped

        block_lines = []
        if matches:
            any_recalls = True
            block_lines.append(f"⚠️ {item} — {len(matches)} active recall(s)")
            shown = matches[:5]
            for m in shown:
                term_status = m.get("terminated", "unknown")
                term_label = (
                    "Terminated" if term_status == "Yes"
                    else "Not terminated" if term_status == "No"
                    else "Terminated status unknown"
                )
                block_lines.append(
                    f"   [{m['source']}] {m['brand']} — {m['reason']} "
                    f"({m['date']}) — {term_label}"
                )
            if len(matches) > len(shown):
                block_lines.append(f"   …and {len(matches) - len(shown)} more not shown")
        else:
            block_lines.append(f"✅ {item} — clear")
        item_blocks.append("\n".join(block_lines))

    body = "\n\n".join(item_blocks)

    if not any_recalls:
        footer = "Nothing on this list has an active CA recall right now."
    else:
        footer = (
            "A Terminated Recall is a recall where the FDA has determined "
            "that all reasonable efforts have been made to remove or correct "
            "the violative product in accordance with the recall strategy, "
            "and proper disposition has been made according to the degree of "
            "hazard. Recalls that are not indicated as being terminated are "
            "either ongoing or completed."
        )

    return f"{header}\n\n{body}\n\n{footer}\n\nMore info: {FDA_PRESS_URL}"


# ---------- diagnostics ----------

def debug_check(term):
    """Runs each source and reports raw status/counts instead of a clean
    yes/no answer — lets us see what's actually happening on the server
    (blocked request? empty page? parsing miss?) instead of guessing."""
    lines = [f"🔧 Debug v7 for: {term}\n"]

    # openFDA
    try:
        params = {
            "search": f'product_description:"{term}"+AND+status:"Ongoing"',
            "sort": "report_date:desc",
            "limit": 25,
        }
        r = requests.get(FDA_URL, params=params, timeout=15)
        body_preview = r.text[:150].replace("\n", " ")
        lines.append(f"[openFDA API] status={r.status_code}")
        if r.status_code == 200:
            n = len(r.json().get("results", []))
            lines.append(f"  raw results (before CA filter): {n}")
        else:
            lines.append(f"  body: {body_preview}")
    except requests.RequestException as e:
        lines.append(f"[openFDA API] request failed: {e}")

    # USDA FSIS API
    try:
        r = requests.get(
            USDA_URL,
            params={"field_product_items_value": term, "field_active_notice": "True"},
            headers=HEADERS,
            timeout=15,
        )
        lines.append(f"[USDA API] status={r.status_code}")
        if r.status_code == 200:
            data = r.json()
            n = len(data) if isinstance(data, list) else "not a list"
            lines.append(f"  raw results: {n}")
        else:
            lines.append(f"  body: {r.text[:150]}")
    except requests.RequestException as e:
        lines.append(f"[USDA API] request failed: {e}")

    # FDA press-release XLSX export (full dataset, no pagination needed)
    try:
        r = requests.get(FDA_XLSX_URL, headers=HEADERS, timeout=30)
        lines.append(f"[FDA XLSX export] status={r.status_code}, bytes={len(r.content)}")
        if r.status_code == 200:
            wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            lines.append(f"  total rows (incl. header): {len(rows)}")
            if rows:
                lines.append(f"  header: {rows[0]}")
                header_norm = [str(h).strip().lower().replace("-", " ") if h else "" for h in rows[0]]

                def col_idx(fragment):
                    for i, h in enumerate(header_norm):
                        if fragment in h:
                            return i
                    return None

                resolved = {
                    "date": col_idx("date"),
                    "brand": col_idx("brand"),
                    "product": col_idx("product description"),
                    "type": col_idx("product type"),
                    "reason": col_idx("reason"),
                    "company": col_idx("company"),
                }
                lines.append(f"  resolved column indices: {resolved}")
                if any(v is None for v in resolved.values()):
                    lines.append("  ⚠️ one or more columns failed to match!")
            if len(rows) > 1:
                lines.append(f"  first data row: {rows[1]}")
                lines.append(f"  last data row: {rows[-1]}")
    except Exception as e:
        lines.append(f"[FDA XLSX export] failed: {e}")

    # FSIS press page
    try:
        r = requests.get(FSIS_PRESS_URL, headers=HEADERS, timeout=15)
        lines.append(f"[FSIS press page] status={r.status_code}, body length={len(r.text)}")
        soup = BeautifulSoup(r.text, "html.parser")
        links = soup.find_all("a")
        lines.append(f"  links found: {len(links)}")
    except requests.RequestException as e:
        lines.append(f"[FSIS press page] request failed: {e}")

    return "\n".join(lines)


# ---------- Telegram webhook ----------

def send_message(chat_id, text):
    requests.post(
        f"{TELEGRAM_API}/sendMessage",
        json={"chat_id": chat_id, "text": text},
        timeout=10,
    )


@app.route("/webhook", methods=["POST"])
def webhook():
    update = request.get_json(force=True, silent=True) or {}
    message = update.get("message") or update.get("edited_message")
    if not message or "text" not in message:
        return "ok"

    chat_id = message["chat"]["id"]
    text = message["text"]

    if text.strip().lower() in ("/start", "/help"):
        send_message(
            chat_id,
            "Send me your grocery list (one item per line) and I'll check it "
            "against active FDA/USDA recalls for California.",
        )
        return "ok"

    if text.strip().lower().startswith("/debug"):
        term = text.strip()[6:].strip()
        if not term:
            send_message(chat_id, "Usage: /debug blueberries")
            return "ok"
        send_message(chat_id, debug_check(term))
        return "ok"

    items = parse_list(text)
    send_message(chat_id, build_reply(items))
    return "ok"


@app.route("/", methods=["GET"])
def health():
    return "Recall bot is running. Visit /check to use it in a browser."


WEB_FORM_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Grocery Recall Check</title>
  <style>
    body { font-family: -apple-system, sans-serif; max-width: 600px; margin: 30px auto; padding: 0 16px; color: #222; }
    h1 { font-size: 22px; }
    textarea { width: 100%; min-height: 120px; font-size: 15px; padding: 10px; box-sizing: border-box; }
    button { margin-top: 10px; padding: 10px 20px; font-size: 15px; background: #222; color: #fff; border: none; border-radius: 4px; cursor: pointer; }
    pre { white-space: pre-wrap; background: #f5f5f0; padding: 14px; border-radius: 4px; font-size: 14px; margin-top: 20px; }
    .hint { color: #666; font-size: 13px; }
  </style>
</head>
<body>
  <h1>🛒 Grocery Recall Check</h1>
  <p class="hint">Paste your list, one item per line. Checked against active FDA/USDA recalls for California / nationwide.</p>
  <form method="POST">
    <textarea name="list" placeholder="blueberries&#10;ground beef&#10;eggs">{{ list_text }}</textarea><br>
    <button type="submit">Check list</button>
  </form>
  {% if result %}
  <pre>{{ result }}</pre>
  {% endif %}
</body>
</html>
"""


@app.route("/check", methods=["GET", "POST"])
def web_check():
    result = None
    list_text = ""
    if request.method == "POST":
        list_text = request.form.get("list", "")
        items = parse_list(list_text)
        result = build_reply(items)
    return render_template_string(WEB_FORM_HTML, result=result, list_text=list_text)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
